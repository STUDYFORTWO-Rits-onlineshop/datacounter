import openpyxl
import os
class BookInfo:
    def __init__(self,title,size,url,count=1,author=None,moreauther=None,publisher=None,page=None,List_price=None,price=None):
        self.title=title
        self.count=count
        self.author=author
        self.moreauther=moreauther
        self.publisher=publisher
        self.size=size
        self.url=url
        self.price=price
        self.List_price=List_price
        self.page=page

os.getcwd()
#ファイルの読み込み
wb1=openpyxl.load_workbook('2020.4.1 在庫リスト.xlsx')
#シートの読み込み
sheet1=wb1["2020.4.1 在庫リスト"]
#変更先の読み込み
sheet2=wb1["miyasuku"]
#rowは行、columnは列
print(sheet1.cell(row=1, column=2).value,sheet1.cell(row=1, column=13).value,sheet1.cell(row=1, column=24).value)
retu=1
dic_list={}
#BookinfoのISBNをキー、それ以外の情報で作成したBookInfoインスタンスを作成して、valueとして対応つけて
#それを１列目が空白である場所まで繰り返す
while sheet1.cell(row=retu, column=1).value!=None:
    if sheet1.cell(row=retu, column=1).value not in dic_list:
        dic_list[sheet1.cell(row=retu, column=1).value]=BookInfo(sheet1.cell(row=retu, column=2).value,sheet1.cell(row=retu, column=16).value,sheet1.cell(row=retu, column=17).value,1,
        sheet1.cell(row=retu, column=6).value,sheet1.cell(row=retu, column=10).value,sheet1.cell(row=retu, column=13).value,
        sheet1.cell(row=retu, column=36).value,sheet1.cell(row=retu, column=24).value)
    else:
        #もうすでに要素が作られていた場合、数(count)を＋１する
        dic_list[sheet1.cell(row=retu, column=1).value].count+=1
    #次の列に行く
    retu+=1
retu=1
for i in dic_list:
    #それぞれを何行目に置くかを要素として受け取る
    dic={dic_list[i].title:1,dic_list[i].size:8,dic_list[i].url:10,dic_list[i].count:7,dic_list[i].author:2,dic_list[i].moreauther:3,
    dic_list[i].publisher:6,dic_list[i].page:9,dic_list[i].List_price:5}
    sheet2.cell(row=retu, column=11,value=i)
    for j,k in dic.items():
        #実際にセルに書き出す
        sheet2.cell(row=retu,column=k,value=j)
    retu+=1
#セーブさせて終了
wb1.save('2020.4.1 在庫リスト.xlsx')
